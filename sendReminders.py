# Send emails based on payment status in excel

import openpyxl
import smtplib
import sys

# open spreadsheet ans get due status

wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
lastCol = sheet.max_column
latestMonth = sheet.cell[1][lastCol].value

# check members paid status
unpaidMembers = {}

for i in range(2, sheet.max_row + 1):
    payment = sheet.cell[i][lastCol].value
    if payment != 'paid':
        name = sheet.cell[i][1].value
        email = sheet.cell[i][2].value
        unpaidMembers[name] = email

# Log in to email
# Placeholder information entered
smtpObj = smtplib.SMTP('smtp.example.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login('example@example.com', sys.argv[1])

# send out emails
for name, email in unpaidMembers.items():
    body = f'Subject: {latestMonth} dues unpaid.\nDear {name},\nRecords show that yo' \
           f'u have not paid dues for {latestMonth}. Please make this payment as soon as possible. Thank you!'
    print(f'Sending email to {email}...')
    sendMail = smtpObj.sendmail('example@example.com', email, body)  # placeholder

    if sendMail != {}:
        print(f'There was a problem sending email to {email}: {sendMail}')
    smtpObj.quit()
